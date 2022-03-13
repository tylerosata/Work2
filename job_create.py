import openpyxl
import os
import PyPDF2
import re
import glob

files = glob.glob('C:\\Users\\otyle\\spam\\new jobs\\*.pdf')
i = 0
bags = dict()
today = str(date.today())
now = int(today[8:10])
dueDate = now + 7
mainPanel = '_M'
combined = '_C'
gusset = '_G'
projectTitle = " ".join(map(str,sys.argv[1:]))
browser = webdriver.Chrome()
action = webdriver.ActionChains(browser) # is this being used?

def firstBag():
    if "BB" not in details[2]:
        panel = mainPanel
        serviceOrder(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        copyINproject()
        gusset(details[0],details[3])
        savejob()
    else:
        panel = combined
        serviceOrder(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        
def nextBags():
    if "BB" not in details[2]:
        panel = mainPanel
        copyInproject()
        newBag(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        copyINproject()
        gusset(details[0],details[3])
        savejob()
    else:
        panel = combined
        copyInproject()
        newBag(details[0],details[1],details[2],details[3],details[4],panel)
        datepicker()
        savejob()
        
while i <= len(files)-1:
        # gets the mt number and bag size from pdf name
    MTnumber = int(re.search(r'\d{6}',files[i]).group())
    bagSize = re.search(r'\d?\.?\d?#',files[i]).group()
        #get the dieline, UPC, and description(name) from tracker based on MT number from pdf
    wb = openpyxl.load_workbook('timeline_tracker2.xlsx')
    sheet = wb['Timeline']
    for rowNum in range(4, sheet.max_row):
        mtNumbers = sheet.cell(row=rowNum, column = 17).value
        if MTnumber == mtNumbers:
            dieline = sheet.cell(row=rowNum, column=14).value
            description = sheet.cell(row=rowNum, column=13).value 
            UPC = sheet.cell(row=rowNum, column=12).value
    wb.close()
    bags['bag_%d'% i] = [MTnumber, bagSize, dieline, description, UPC]
    i +=1
# print(bags)

########################################################################################
### Create project and create first job - it will already have the BB/GMI information###
########################################################################################

hubxlogin()
createProject()
addJob()

for i in bags:
    details = bags.get(i)
    if i <= 0:
        firstBag()
    else:
        nextBags()
